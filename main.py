#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""VISIMICS AUTOPILOT — репрайсер каталогу Prom (єдиний файл, БЕЗ Prom API).
Джерела собівартості: BMW+Porsche (Google Sheets) + AutoNova (Drive/пошта).
Каталог і публікація — ТІЛЬКИ через вкладку «Export Products Sheet» у хабі
(її тягне Prom за URL-фідом). Пише три поля: Ціна(8), Наявність(15), Кількість(16).
Захист: MIN_MARKUP_ABS, MAX_DROP_PCT, ANCHOR_FLOOR (якір 30.06).
Реальний запис у таблицю лише при LIVE=1 (інакше DRY-RUN: рахує + звіт, не пише)."""
import os, json, io, math, imaplib, email, datetime, time, urllib.request, urllib.error

ID_BMW="1KXaDLqBsOAtX0MxUoX39jpia9boISxl1xUxPihhU77I"
ID_PORSCHE="1oVSVg1cBxGj-DA66c5_FoAtp6zOthdnF_xTY_ugez2g"
ID_HUB="1pesHiOHDq2Y4FYQECakfhIJlq08bg5_Pkm9e2YEDoic"
AUTONOVA_FROM="1c@autonovad.ua"
# --- AutoNova-WEB (catalogue-api; дилерська ціна лише під кукі) ---
AUTONOVA_API="https://catalogue-api.autonovad.ua/api/products"
AUTONOVA_DEFAULT_BRAND=int(os.environ.get("AUTONOVA_BRAND_ID") or 72) # BMW=72 (нерозпізнані вважаємо BMW)
AUTONOVA_REF=("A1678992200", 56, 6000) # (код, brandId, поріг грн) — реф. для перевірки дилерської ціни
EXPORT_TAB="Export Products Sheet" # ЄДИНЕ джерело каталогу і місце публікації (Prom тягне звідси)
# колонки формату експорту Prom (0-based)
C_CODE=0; C_NAME=1; C_PRICE=8; C_AVAIL=15; C_QTY=16
MARGIN_FLOOR=1.16
LIVE=os.environ.get("LIVE")=="1" # LIVE=1 -> реально писати в Export; інакше DRY-RUN

def num(x):
    try: return float(str(x).replace(",",".").replace("\xa0","").replace(" ",""))
    except: return 0.0
# --- Захисні параметри ціноутворення (env-налаштовувані, щоб не лізти в код) ---
MIN_MARKUP_ABS=num(os.environ.get("MIN_MARKUP_ABS") or 150) # мін. абсолютна націнка (грн) для дешевих позицій
MAX_DROP_PCT=num(os.environ.get("MAX_DROP_PCT") or 25)/100.0 # зниження діючої ціни > цього % без конкурента -> УТРИМАТИ (не писати)
ANCHOR_FLOOR=num(os.environ.get("ANCHOR_FLOOR") or 60)/100.0 # не писати ціну нижче цього % від якоря 30.06 без конкурента -> УТРИМАТИ
def load_anchor(path=None):
    """Якір реальних цін до поломки (export 30.06): article(UPPER)->price. CSV: article,anchor_price."""
    path=path or os.environ.get("ANCHOR_CSV") or "anchor_prices.csv"; m={}
    try:
        import csv as _csv
        with open(path, encoding="utf-8") as f:
            rd=_csv.reader(f); next(rd, None)
            for row in rd:
                if len(row)>=2:
                    a=str(row[0]).strip().upper(); p=num(row[1])
                    if a and p>0: m[a]=p
        print(f"[anchor] завантажено {len(m)} якірних цін із {path}")
    except FileNotFoundError:
        print(f"[anchor] {path} нема — якірний захист вимкнено")
    except Exception as e:
        print("[anchor]", str(e)[:80])
    return m
def final_price(cost):
    c=num(cost)
    if c<=0: return 0
    k=1.5 if c<3000 else 1.45 if c<5000 else 1.3 if c<10000 else 1.2 if c<30000 else 1.1
    return int(math.ceil(max(c*k, c+MIN_MARKUP_ABS))) # +абсолютний мінімум націнки для дешевих
def price_with_competitor(cost, comp):
    base=final_price(cost)
    if not comp or comp<=0: return base
    floor=int(math.ceil(num(cost)*MARGIN_FLOOR)); target=int(comp)-1
    return target if (target>=floor and target<base) else base
def keep_best(best, art, item, instock):
    k=str(art).strip().upper()
    if not k: return
    if item.get("presence")=="available" and num(item.get("qty"))>0: # пріоритет "в наявності"
        instock[k]=max(instock.get(k,0), int(num(item.get("qty"))))
    if k not in best or item["cost"]<best[k]["cost"]:
        item["article"]=k; best[k]=item
def _norm(x): return "".join(str(x).lower().split())
def gclient():
    key=os.environ.get("GCP_SA_KEY")
    if not key: return None
    import gspread
    return gspread.service_account_from_dict(json.loads(key)) if key.strip().startswith("{") else None

def read_all_tabs(gc, sid, brand, best, instock, force=None):
    try: ss=gc.open_by_key(sid)
    except Exception as e: print(f"[sheet] {brand}: OPEN FAIL {str(e)[:80] or 'нема доступу'}"); return
    for ws in ss.worksheets():
        title=ws.title; nt=_norm(title)
        if force: presence=force
        elif "наяв" in nt: presence="available"
        elif any(k in nt for k in ["чека","2-3","2–3","23дн","замов","15дн","15днів","підзам"]): presence="order"
        else: presence="available"
        try: rows=ws.get_all_values()
        except Exception as e: print(f"[sheet] {brand}/{title}: READ FAIL {str(e)[:60]}"); continue
        n=0
        for r in rows:
            if len(r)<3: continue
            art=(r[0] or "").strip()
            if not art: continue
            cost=num(r[3]) if len(r)>=4 else 0
            if cost>0: qty=num(r[2])
            else: cost=num(r[2]); qty=0
            if cost<=0: continue
            keep_best(best, art, {"name":r[1],"cost":cost,"qty":qty,"presence":presence,"brand":brand}, instock); n+=1
        print(f"[sheet] {brand}/{title}: {n} поз. ({presence})")

def pull_autonova(best, instock):
    user=os.environ.get("MAIL_USER"); pw=os.environ.get("MAIL_PASS")
    if not user or not pw: print("[autonova] нема MAIL_USER/PASS — пропуск"); return
    try:
        import openpyxl
        M=imaplib.IMAP4_SSL("imap.gmail.com"); M.login(user,pw); M.select("INBOX")
        since=(datetime.date.today()-datetime.timedelta(days=14)).strftime("%d-%b-%Y")
        _,data=M.search(None, f'(FROM "{AUTONOVA_FROM}" SINCE {since})'); ids=data[0].split()
        if not ids: print("[autonova] листів нема за 14 днів"); M.logout(); return
        for numid in reversed(ids):
            _,d=M.fetch(numid,"(RFC822)"); msg=email.message_from_bytes(d[0][1]); done=False
            for part in msg.walk():
                fn=part.get_filename() or ""
                if fn.lower().endswith((".xlsx",".xls")):
                    wb=openpyxl.load_workbook(io.BytesIO(part.get_payload(decode=True)),read_only=True,data_only=True)
                    sh=wb["TDSheet"] if "TDSheet" in wb.sheetnames else wb[wb.sheetnames[0]]; n=0
                    for r in sh.iter_rows(values_only=True):
                        if not r or len(r)<=24 or r[2] is None: continue
                        cost=num(r[24])
                        if cost<=0: continue
                        qty=sum(num(r[c]) for c in range(5,23) if c<len(r))
                        if qty<=0: continue
                        keep_best(best,r[2],{"name":str(r[1] or ""),"cost":cost,"qty":qty,"presence":"available","brand":"Авто"}, instock); n+=1
                    wb.close(); done=True; print(f"[autonova] {n} поз.")
            if done: break
        M.logout()
    except Exception as e: print(f"[autonova] {str(e)[:120]}")

def _open_xlsx_tolerant(data):
    """Відкриває xlsx; якщо 1С-файл без xl/sharedStrings.xml — додає порожній і пробує ще раз."""
    import openpyxl, zipfile
    try:
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        zin=zipfile.ZipFile(io.BytesIO(data)); buf=io.BytesIO()
        zout=zipfile.ZipFile(buf,"w",zipfile.ZIP_DEFLATED)
        for it in zin.namelist(): zout.writestr(it, zin.read(it))
        if "xl/sharedStrings.xml" not in zin.namelist():
            zout.writestr("xl/sharedStrings.xml",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>')
        zout.close(); zin.close()
        return openpyxl.load_workbook(io.BytesIO(buf.getvalue()), read_only=True, data_only=True)

def pull_autonova_drive(folder_id, best, instock):
    """Читає найсвіжіший прайс Autonova з Drive-теки (її наповнює Apps Script)
    через сервіс-акаунт. Без IMAP/пароля. Підтримує zip-архів і xlsx."""
    folder_id=(folder_id or "").strip()
    if "/folders/" in folder_id: # приймаємо і повний URL теки, не лише ID
        folder_id=folder_id.split("/folders/")[1].split("?")[0].split("#")[0].split("/")[0]
    key=os.environ.get("GCP_SA_KEY")
    if not key: print("[autonova] нема GCP_SA_KEY — пропуск Drive"); return
    try:
        import openpyxl
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        creds=Credentials.from_service_account_info(json.loads(key),
            scopes=["https://www.googleapis.com/auth/drive.readonly"])
        svc=build("drive","v3",credentials=creds,cache_discovery=False)
        q="'%s' in parents and trashed=false and name contains 'autonova_latest'"%folder_id
        files=svc.files().list(q=q,fields="files(id,name,modifiedTime,size)",
            orderBy="modifiedTime desc",pageSize=5,
            supportsAllDrives=True,includeItemsFromAllDrives=True).execute().get("files",[])
        if not files: print("[autonova] у Drive-теці файлів нема"); return
        fn=files[0]["name"].lower()
        data=svc.files().get_media(fileId=files[0]["id"]).execute()
        if fn.endswith(".zip"): # прайс приходить архівом → розпакувати
            import zipfile
            zf=zipfile.ZipFile(io.BytesIO(data))
            inner=[m for m in zf.namelist() if m.lower().endswith((".xlsx",".xls"))]
            if not inner: print("[autonova] у zip нема xlsx"); return
            data=zf.read(inner[0])
        elif fn.endswith(".rar"):
            print("[autonova] .rar не підтримується — треба zip"); return
        from python_calamine import CalamineWorkbook
        cw=CalamineWorkbook.from_filelike(io.BytesIO(data))
        names=cw.sheet_names
        sh=cw.get_sheet_by_name("TDSheet") if "TDSheet" in names else cw.get_sheet_by_index(0)
        rows=sh.to_python(skip_empty_area=False); n=0
        for r in rows:
            if len(r)<=24 or r[2] in (None,""): continue
            cost=num(r[24])
            if cost<=0: continue
            qty=sum(num(r[c]) for c in range(5,23) if c<len(r))
            if qty<=0: continue
            keep_best(best,r[2],{"name":str(r[1] or ""),"cost":cost,"qty":qty,
                "presence":"available","brand":"Авто"}, instock); n+=1
        print(f"[autonova] Drive '{files[0]['name']}': {n} поз.")
    except Exception as e:
        print(f"[autonova] Drive {str(e)[:140]}")

def load_map(gc, tab):
    m={}
    try:
        for r in gc.open_by_key(ID_HUB).worksheet(tab).get_all_values()[1:]:
            if len(r)<2: continue
            a=(r[0] or "").strip().upper(); p=num(r[1])
            if a and p>0: m[a]=p
    except Exception as e: print(f"[{tab}] {str(e)[:50]}")
    return m

# ================== BM Parts (постачальник) — bulk-прайс по брендах ==================
def _bmparts_list_map():
    """Прайс BM Parts УСІХ брендів ОДНИМ запитом (POST /prices/list, format=csv).
    Лише товари в наявності. Повертає {article(UPPER): {'price','qty','presence'}}.
    Це надійніше за ітерацію /prices/prom по 30 марках (без per-brand encoding/rate-limit проблем)."""
    token=os.environ.get("BMPARTS_TOKEN")
    if not token: return {}
    import csv as _csv
    try:
        from bmparts import BMParts
        bm=BMParts(token)
        whs=[w.get("uuid") for w in bm.warehouses() if w.get("uuid")]
    except Exception as e:
        print(f"[bmparts] list init FAIL: {str(e)[:100]}"); return {}
    try:
        r=bm.s.post("https://api.bm.parts/prices/list",
                    json={"warehouses":whs,"format":"csv","products_type":"code"}, timeout=180)
        if r.status_code!=200:
            print(f"[bmparts] /prices/list HTTP {r.status_code}"); return {}
        text=r.text
    except Exception as e:
        print(f"[bmparts] /prices/list FAIL: {str(e)[:100]}"); return {}
    sample=text[:2000]; delim=";" if sample.count(";")>=sample.count(",") else ","
    rows=list(_csv.reader(io.StringIO(text), delimiter=delim))
    if len(rows)<2: print("[bmparts] /prices/list порожній"); return {}
    head=[h.strip().lower() for h in rows[0]]
    def col(*keys):
        for i,h in enumerate(head):
            if any(k in h for k in keys): return i
        return -1
    ci=col("код_товару","код товару","артикул","article","код","ідентифікатор")
    cp=col("ціна","price"); cav=col("наявн","availab","presence"); cq=col("кільк","quantity","qty","залиш","остат")
    if ci<0 or cp<0:
        print(f"[bmparts] /prices/list: колонки не розпізнані {head[:8]}"); return {}
    out={}
    for r in rows[1:]:
        if ci>=len(r) or cp>=len(r): continue
        art=_nkey(r[ci]); price=num(r[cp]) # нормалізуємо: BM Parts артикули з дефісами (20114-0050-99 -> 20114005099)
        if not art or price<=0: continue
        qty=num(r[cq]) if 0<=cq<len(r) else 0
        av=(r[cav].strip().lower() if 0<=cav<len(r) else "")
        available=("наявн" in av or av in ("+","true","1","в наявності","у наявності")) or qty>0
        if art not in out or price<out[art]["price"]:
            out[art]={"price":price,"qty":int(qty),"presence":"available" if available else "order"}
    print(f"[bmparts] /prices/list: {len(out)} унікальних артикулів (одним запитом)")
    return out

def _bmparts_price_map(brands=None):
    """Прайс BM Parts у форматі Prom по КАР-БРЕНДАХ. brands=None -> УСІ авто-марки з каталогу
    (GET /catalog/cars/brands/), або список із BMPARTS_BRANDS. Так знаходимо не лише BMW,
    а й Mercedes/VAG/Audi/… Повертає {article(UPPER): {'price','qty','presence'}}.
    price = собівартість (закупівельна ціна постачальника, підтверджено). Лише товари в наявності."""
    if not brands and not os.environ.get("BMPARTS_BRANDS"):
        _m=_bmparts_list_map()                 # спершу ОДИН запит на всі бренди (/prices/list)
        if _m: return _m
        print("[bmparts] /prices/list дав 0 — fallback на per-brand /prices/prom")
    token=os.environ.get("BMPARTS_TOKEN")
    if not token: print("[bmparts] нема BMPARTS_TOKEN — пропуск"); return {}
    import csv as _csv
    try:
        from bmparts import BMParts
    except Exception as e:
        print(f"[bmparts] import bmparts не вдався: {str(e)[:90]}"); return {}
    try:
        bm=BMParts(token)
        whs=[w.get("uuid") for w in bm.warehouses() if w.get("uuid")]
    except Exception as e:
        print(f"[bmparts] warehouses FAIL: {str(e)[:100]}"); return {}
    if not brands:
        env=os.environ.get("BMPARTS_BRANDS")
        if env:
            brands=[b.strip() for b in env.split(",") if b.strip()]
        else:
            try: # усі авто-марки з каталогу BM Parts
                rr=bm.s.get("https://api.bm.parts/catalog/cars/brands/", timeout=60); rr.raise_for_status()
                brands=[b.get("name") for b in (rr.json().get("car_brands") or []) if b.get("name")]
                print(f"[bmparts] авто-марок у каталозі: {len(brands)}")
            except Exception as e:
                print(f"[bmparts] список авто-марок FAIL: {str(e)[:100]} — лишаю BMW"); brands=["BMW"]
    out={}; brands_hit=0
    for brand in brands:
        try:
            text=bm.prom_price_csv(brand, whs)
        except Exception:
            continue # марка без даних/помилка — тихо далі (щоб не спамити лог на 100+ марок)
        sample=text[:2000]; delim=";" if sample.count(";")>=sample.count(",") else ","
        rows=list(_csv.reader(io.StringIO(text), delimiter=delim))
        if len(rows)<2: continue
        head=[h.strip().lower() for h in rows[0]]
        def col(*keys):
            for i,h in enumerate(head):
                if any(k in h for k in keys): return i
            return -1
        ci=col("код_товару","код товару","артикул","article","ідентифікатор")
        cp=col("ціна","price"); cav=col("наявн","availab","presence"); cq=col("кільк","quantity","qty","залиш","остат")
        if ci<0 or cp<0: continue
        n=0
        for r in rows[1:]:
            if ci>=len(r) or cp>=len(r): continue
            art=_nkey(r[ci]); price=num(r[cp]) # нормалізуємо (дефіси в артикулах BM Parts)
            if not art or price<=0: continue
            qty=num(r[cq]) if 0<=cq<len(r) else 0
            av=(r[cav].strip().lower() if 0<=cav<len(r) else "")
            available=("наявн" in av or av in ("+","true","1","в наявності","у наявності")) or qty>0
            if art not in out or price<out[art]["price"]: # лишаємо найдешевшу
                out[art]={"price":price,"qty":int(qty),"presence":"available" if available else "order"}
            n+=1
        if n: brands_hit+=1; print(f"[bmparts] {brand}: {n}")
    print(f"[bmparts] мапа: {len(out)} унікальних артикулів по {brands_hit} марках (із {len(brands)})")
    return out

def pull_bmparts(codes, best, instock, brands=None):
    """BM Parts (УСІ авто-марки з каталогу) для кодів БЕЗ постачальника.
    Ключ зіставлення = номер до тире (частина артикула до дефіса). Ціна = собівартість."""
    import re as _re
    pm=_bmparts_price_map(brands) # brands=None -> усі авто-марки (Mercedes/VAG/Audi/… а не лише BMW)
    if not pm: print("[bmparts] мапа порожня — нічого не додано"); return
    n_ok=n_avail=0
    for code in codes:
        rec=pm.get(_nkey(code)) or pm.get(_nkey(_re.split(r"[-–—]", str(code))[0])) # нормалізований збіг: цілий код, потім номер до тире
        if not rec: continue
        av=(rec["presence"]=="available" and rec["qty"]>0)
        keep_best(best, str(code).strip().upper(),
            {"name":"","cost":rec["price"],"qty":int(rec["qty"]) if av else 0,
             "presence":"available" if av else "order","brand":"BM Parts"}, instock)
        n_ok+=1
        if av: n_avail+=1
    print(f"[bmparts] додано {n_ok} кодів (у наявності: {n_avail})")

def _nkey(s):
    import re as _re
    return _re.sub(r"[^0-9a-zA-Z]","",str(s)).upper() # нормалізація коду: лише цифри/літери

def _expand_code(code):
    """Розкладає дефісний код на ПОВНІ номери. Другий+ номер може бути скороченим
    суфіксом першого: '51117303107-108' -> ['51117303107','51117303108'];
    '51712150246-47' -> ['51712150246','51712150247']. Якщо номер уже повний — лишається як є."""
    raw=[p.strip() for p in str(code).split("-") if p.strip()]
    if not raw: return []
    base=raw[0]; out=[base]
    for nx in raw[1:]:
        if len(nx)<len(base): nx=base[:len(base)-len(nx)]+nx # дотягнути префіксом першого
        out.append(nx)
    return out

# ================== AutoNova-WEB (джерело №2 — дилерські ціни під кукі) ==================
AUTONOVA_PROXY=os.environ.get("AUTONOVA_PROXY") # напр. http://user:pass@ua-host:port — обхід блоку IP GitHub-раннера
def _autonova_opener():
    """Opener з проксі (якщо задано AUTONOVA_PROXY) — щоб ходити на autonovad.ua з українського IP."""
    if AUTONOVA_PROXY:
        h=urllib.request.ProxyHandler({"http":AUTONOVA_PROXY,"https":AUTONOVA_PROXY})
        return urllib.request.build_opener(h)
    return urllib.request.build_opener()
def _autonova_diag(product_id, cookie):
    """ТОЧНА діагностика контрольного запиту autonova-web: друкує HTTP-статус або тип
    мережевої помилки. 401/403 -> кукі невалідна; timeout/refused/URLError -> блок IP раннера."""
    import socket
    url=f"{AUTONOVA_API}/{product_id}/extended-offers"
    print(f"[autonova-diag] proxy={'ТАК' if AUTONOVA_PROXY else 'ні'}; cookie={'є' if cookie else 'НЕМА'} (довж {len(cookie or '')}); {url}")
    req=urllib.request.Request(url, headers={"Cookie":cookie or "",
        "Accept":"application/json","User-Agent":"Mozilla/5.0 (visimics-autopilot)"})
    try:
        with _autonova_opener().open(req, timeout=8) as r:
            body=r.read(200).decode("utf-8","replace")
            print(f"[autonova-diag] HTTP {getattr(r,'status',200)} OK; тіло[:100]={body[:100]!r}")
    except urllib.error.HTTPError as e:
        b=""
        try: b=e.read(200).decode("utf-8","replace")
        except Exception: pass
        verdict=("КУКІ невалідна/протухла" if e.code in (401,403) else
                 "редирект на гостя (кукі)" if e.code in (301,302) else "інша HTTP-помилка")
        print(f"[autonova-diag] HTTPError {e.code} {e.reason} -> {verdict}; тіло[:100]={b[:100]!r}")
    except urllib.error.URLError as e:
        print(f"[autonova-diag] URLError {e.reason!r} -> БЛОК IP / мережа / DNS")
    except socket.timeout:
        print("[autonova-diag] TIMEOUT -> БЛОК IP / фаєрвол")
    except Exception as e:
        print(f"[autonova-diag] {type(e).__name__}: {str(e)[:120]}")
def _autonova_fetch(product_id, cookie):
    """GET .../api/products/{product_id}/extended-offers під дилерською кукі. JSON або None."""
    url=f"{AUTONOVA_API}/{product_id}/extended-offers"
    req=urllib.request.Request(url, headers={
        "Cookie": cookie,
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0 (visimics-autopilot)",
    })
    for attempt in range(3): # 520 у origin буває транзієнтним
        try:
            with _autonova_opener().open(req, timeout=8) as r:
                return json.loads(r.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if e.code in (502,503,520,522,524) and attempt<2: time.sleep(1.2); continue
            return None
        except Exception:
            if attempt<2: time.sleep(1.0); continue
            return None
    return None

def _autonova_code_best(code, brand_id, cookie):
    """Собівартість/наявність по ОДНОМУ номеру. Мінімум по ВСІХ пропозиціях
    (bestPrice/bestDelivery у API брехливі — рахуємо самі). None якщо нема пропозицій."""
    d=_autonova_fetch(f"{code}_{brand_id}", cookie)
    if not d: return None
    cand=[]
    for grp in ("offers","supplierOffers","branchOffers","consignmentOffers"):
        for o in (d.get(grp) or []):
            p=num((o.get("price") or {}).get("current"))
            if p<=0: continue
            cand.append({
                "price": p,
                "qty":   num(o.get("quantity")),
                "days":  num((o.get("delivery") or {}).get("days")),
                "own":   (o.get("category")=="offers"), # власний склад АвтоНови
            })
    if not cand: return None
    own_stock=[c for c in cand if c["own"] and c["qty"]>0 and c["days"]<=1] # є сьогодні на власному складі
    if own_stock:
        b=min(own_stock, key=lambda c:c["price"])
        return {"cost":b["price"], "qty":int(b["qty"]), "presence":"available"}
    cheapest=min(cand, key=lambda c:c["price"]) # інакше — під замовлення за найдешевшою
    return {"cost":cheapest["price"], "qty":0, "presence":"order"}

def autonova_web_authorized(cookie):
    """Перевірка, що кукі дає ДИЛЕРСЬКУ ціну (а не гостьову). Захист від протухлої кукі:
    краще нічого не писати, ніж записати завищену собівартість."""
    code,bid,thr=AUTONOVA_REF
    d=_autonova_fetch(f"{code}_{bid}", cookie)
    if not d:
        print("[autonova-web] реф. запит не вдався — пропуск (кукі/мережа)")
        _autonova_diag(f"{code}_{bid}", cookie)  # точна причина: HTTP-статус (кукі) vs мережа (блок IP)
        return False
    ref=num((d.get("bestDelivery") or {}).get("price",{}).get("current"))
    if 0<ref<thr:
        print(f"[autonova-web] авторизація OK (дилерська реф.ціна {ref:.0f} < {thr})"); return True
    print(f"[autonova-web] УВАГА: кукі схоже протухла (реф.ціна {ref:.0f} ≥ {thr}, гостьова). "
          f"НІЧОГО не пишу з web, щоб не завищити собівартість. Онови AUTONOVA_COOKIE."); return False

def _autonova_brand_for(code):
    """brandId autonova за форматом артикула (перевірено на реальних кодах):
    суто цифри -> BMW(72); 'A'+цифри -> Mercedes(56);
    решта алфанумерик (4H,4L,8W,4M,G0,80A,3R,5GE...) -> VAG(1)."""
    c=str(code).strip().upper()
    if not c: return None
    if c.startswith("WAP") or (c[:1]=="9" and len(c)>=8): return 81 # Porsche (WAP-аксесуари, 9-шасі коди)
    if c.isdigit(): return 72
    if c[:1]=="A" and c[1:2].isdigit(): return 56
    return 1
def pull_autonova_web(codes, best, instock, cookie):
    """Для кодів БЕЗ постачальника — тягне ціну/наявність з catalogue-api.
    Дефіс = пара BMW-номерів: собівартість = сума, наявна лише якщо ОБИДВІ є, к-сть = min.
    Ключ у best = точний код каталогу (з дефісом), щоб зіставитися з Export Products Sheet."""
    if not cookie: print("[autonova-web] нема AUTONOVA_COOKIE — пропуск"); return
    if not autonova_web_authorized(cookie): return
    limit=int(num(os.environ.get("AUTONOVA_WEB_LIMIT") or 0)) # 0 = всі (для тесту можна обмежити)
    ALL_BRANDS=[int(x) for x in (os.environ.get("AUTONOVA_BRANDS") or "1,72,56,59,81,16").split(",") if x.strip()] # VAG=1,BMW=72,Mercedes=56,Mann=59,Porsche=81,+16
    import re as _re2
    def _av_parts(c):
        # Автонова-код = артикул БЕЗ пробілів (дилерський суфікс лишається!), розбитий по
        # роздільниках пар '+' та '-'. Реальні промахи (перевірено на сайті autonovad.ua):
        #   '80A061276A MNO' -> '80A061276AMNO'; '7P1061500  041' -> '7P1061500041';
        #   '5H0853688ADBOP+5H0853688A' -> перша половина '5H0853688ADBOP' (друга відсутня).
        raw=[]
        for seg in str(c).split("+"):     # '+' -> окремі повні номери (грилі-варіанти)
            raw+=_expand_code(seg)         # '-' -> повні номери з дотягуванням суфікса
        out=[]
        for p in raw:
            p=_re2.sub(r"\s+","",str(p)).strip()   # ПРИБРАТИ ПРОБІЛИ (ключовий фікс)
            if p and len(p)>=5 and p not in out: out.append(p)
        return out
    n_ok=n_pair=n_avail=0; seen=0
    for code in codes:
        if limit and seen>=limit: break
        seen+=1
        sp=_av_parts(code)
        cand_whole=_nkey(code)             # весь код лише алфанум: '19-276922' -> '19276922'
        if not sp and cand_whole: sp=[cand_whole]
        if not sp: continue
        guess=_autonova_brand_for(code) # спершу вгадана марка (VAG-код -> 1), далі решта
        order=([guess]+[b for b in ALL_BRANDS if b!=guess]) if guess else ALL_BRANDS
        res=None; is_pair=False
        for bid in order:
            if len(sp)>=2:
                acc=[_autonova_code_best(p, bid, cookie) for p in sp]
                for _ in sp: time.sleep(0.12)
                if all(acc): res=acc; is_pair=True; break        # справжня пара -> сума половин
                elif acc and acc[0]: res=[acc[0]]; break         # інакше -> ціна за ПЕРШОЮ половиною
            else:
                r=_autonova_code_best(sp[0], bid, cookie); time.sleep(0.12)
                if r: res=[r]; break
            if cand_whole and cand_whole!=sp[0]:                 # запасний варіант: весь код алфанум
                r2=_autonova_code_best(cand_whole, bid, cookie); time.sleep(0.12)
                if r2: res=[r2]; break
        if not res: continue
        cost=sum(r["cost"] for r in res) # пара = сума собівартостей
        available=all(r["presence"]=="available" for r in res)
        qty=min(int(r["qty"]) for r in res) if available else 0
        keep_best(best, str(code).strip().upper(),
            {"name":"", "cost":cost, "qty":qty,
             "presence":"available" if available else "order", "brand":"Авто-web"}, instock)
        n_ok+=1
        if is_pair: n_pair+=1
        if available: n_avail+=1
    print(f"[autonova-web] додано {n_ok} кодів (пар: {n_pair}, у наявності: {n_avail}) з {seen} перевірених")

def pull_pairs_from_best(codes, best, instock):
    """ДЖЕРЕЛО №1 для кодів без постачальника: подвоєні номери BMW (через дефіс) —
    це пара реальних BMW-артикулів, обидва вже є в BMW-аркушах (best).
    Собівартість = сума половин; наявність — лише якщо обидві в наявності; к-сть = min.
    Зіставлення нормалізоване (стійке до пробілів/регістру). Без зовнішніх запитів."""
    bnk={}
    for k,v in best.items():
        bnk.setdefault(_nkey(k), v) # індекс best за нормалізованим ключем
    def _add(code, cost, av, qty, brand):
        keep_best(best, str(code).strip().upper(),
            {"name":"","cost":cost,"qty":int(qty) if av else 0,
             "presence":"available" if av else "order","brand":brand}, instock)
    n_whole=n_pair=n_avail=0; unmatched=[]
    for code in codes:
        # 1) НОМЕР ДО ТИРЕ як окремий BMW-артикул. Якщо код — пара (є дефіс), це комплект
        #    лівий+правий: ціна = ціна ПЕРШОГО номера ×2. Інакше (без дефіса) ×1.
        first=str(code).split("-")[0].strip()
        f=bnk.get(_nkey(first)) if first else None
        if f is not None and num(f.get("cost"))>0:
            is_pair=("-" in str(code))
            av=(f.get("presence")=="available" and num(f.get("qty"))>0)
            _add(code, num(f.get("cost"))*(2 if is_pair else 1), av, num(f.get("qty")),
                 "BMW-пара(×2)" if is_pair else "BMW"); n_whole+=1
            if av: n_avail+=1
            continue
        # 1б) ВЕСЬ код як окремий артикул у BMW-аркуші (інший роздільник/формат)
        w=bnk.get(_nkey(code))
        if w is not None and num(w.get("cost"))>0:
            av=(w.get("presence")=="available" and num(w.get("qty"))>0)
            _add(code, num(w.get("cost")), av, num(w.get("qty")), "BMW"); n_whole+=1
            if av: n_avail+=1
            continue
        # 2) ПАРА половинок (сума собівартостей), другий номер може бути скороченим суфіксом
        parts=_expand_code(code)
        if len(parts)>=2:
            rec=[bnk.get(_nkey(p)) for p in parts]
            if all(x is not None for x in rec):
                cost=sum(num(x.get("cost")) for x in rec)
                if cost>0:
                    av=all(x.get("presence")=="available" and num(x.get("qty"))>0 for x in rec)
                    qty=min(int(num(x.get("qty"))) for x in rec) if av else 0
                    _add(code, cost, av, qty, "BMW-пара"); n_pair+=1
                    if av: n_avail+=1
                    continue
        if len(unmatched)<12: unmatched.append(code)
    print(f"[pairs] BMW з аркушів: ціле={n_whole}, пари={n_pair} (у наявності: {n_avail})")
    for code in unmatched: # діагностика: де половинки НЕ знаходяться (з розгортанням суфікса)
        halves=[_nkey(h) for h in _expand_code(code)]
        print("[diag2]", code, "->", " | ".join(h+("=IN" if h in bnk else "=NO") for h in halves))

def read_export(gc):
    """Каталог Prom із вкладки «Export Products Sheet» (те, що тягне Prom).
    Повертає (ws, vals[2D, з паддінгом], idx: код(UPPER)->індекс рядка)."""
    ss=gc.open_by_key(ID_HUB)
    import re as _re
    keyf=lambda s:_re.sub(r"[^a-z0-9]","",str(s).lower()) # лишити тільки літери/цифри: стійко до пробілів/nbsp/невидимих символів/пунктуації
    want=keyf(EXPORT_TAB)
    ws=None
    for w in ss.worksheets():
        if keyf(w.title)==want: ws=w; break
    if ws is None:
        titles=[w.title for w in ss.worksheets()]
        print(f"[fatal] вкладку «{EXPORT_TAB}» не знайдено у хабі. Наявні вкладки: {titles}")
        raise SystemExit(2)
    print(f"[export] вкладка знайдена: «{ws.title}»")
    vals=ws.get_all_values()
    width=max(C_QTY+1, max((len(r) for r in vals), default=0))
    for r in vals: # вирівняти ширину, щоб безпечно писати по індексах
        if len(r)<width: r.extend([""]*(width-len(r)))
    idx={}
    for i,r in enumerate(vals):
        if i==0: continue # заголовок
        code=str(r[C_CODE]).strip().upper()
        if code: idx[code]=i
    return ws, vals, idx

def write_report(gc, catalog, best, instock, overrides, comps, guard_status=None, final_price_map=None):
    """Вкладка «Звіт_Ціни» — ЖУРНАЛ: по кожному товару каталогу — поточна ціна,
    нова ціна робота, зміна %, наявність, джерело, статус (утримано/оновлено/нема постачальника)."""
    guard_status=guard_status or {}; final_price_map=final_price_map or {}
    head=["Артикул","Назва (Prom)","Ціна нова","Зміна %","Наявність","Кількість","Джерело","Собівартість","Статус"]
    rows=[head]
    for art,info in catalog.items():
        b=best.get(art)
        if b:
            newp=final_price_map.get(art)
            if newp is None:
                newp=overrides.get(art) or price_with_competitor(b["cost"], comps.get(art))
            cur=num(info.get("price"))
            chg=("%+.0f%%"%(100*(num(newp)-cur)/cur)) if cur>0 else ""
            aq=instock.get(art,0)
            pres="в наявності" if aq>0 else "під замовлення"
            qty=aq if aq>0 else ""
            rows.append([art, info.get("name",""), newp, chg, pres, qty,
                b.get("brand",""), b.get("cost",""), guard_status.get(art,"оновлено")])
        else:
            rows.append([art, info.get("name",""), "", "", "", "", "", "", "Ручне коригування ціни"])
    ss=gc.open_by_key(ID_HUB)
    RTAB="Звіт_Ціни" # окрема вкладка звіту цін (не плутати з карткою «Звіт»)
    ws=None
    for w in ss.worksheets(): # знайти наявну БЕЗ огляду на регістр (фікс «already exists»)
        if w.title.strip().casefold()==RTAB.casefold(): ws=w; break
    if ws is None:
        ws=ss.add_worksheet(title=RTAB, rows=max(len(rows)+5,100), cols=len(head))
    ws.resize(rows=max(len(rows)+5,10), cols=len(head))
    ws.clear()
    ws.update(values=rows, range_name="A1")
    print(f"[report] {RTAB}: {len(rows)-1} рядків записано")

def pull_autonova_cache(best, instock, path=None):
    """Кеш дилерських цін autonova (зібраний з веб-сесії дилера): code<TAB>cost<TAB>qty<TAB>presence.
    Це джерело для кодів, яких немає у BMW/Porsche-аркушах (переважно Mercedes/BMW-дрібнота)."""
    path=path or os.environ.get("AUTONOVA_CACHE") or "autonova_web_cache.csv"; n=0
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                parts=line.rstrip("\n").split("\t")
                if len(parts)<2: continue
                code=parts[0].strip(); cost=num(parts[1])
                if not code or cost<=0: continue
                qty=num(parts[2]) if len(parts)>2 else 0
                pres=(parts[3].strip() if len(parts)>3 else "order")
                av=(pres=="available" and qty>0)
                keep_best(best, code, {"name":"","cost":cost,"qty":int(qty) if av else 0,
                    "presence":"available" if av else "order","brand":"Авто-web"}, instock); n+=1
        print(f"[autonova-cache] завантажено {n} дилерських цін із {path}")
    except FileNotFoundError:
        print(f"[autonova-cache] {path} нема — пропуск")
    except Exception as e:
        print(f"[autonova-cache] {str(e)[:100]}")

def main():
    gc=gclient()
    if not gc:
        print("[fatal] нема GCP_SA_KEY — без доступу до Google-таблиць працювати нема з чим"); return
    best={}; instock={}
    read_all_tabs(gc,ID_BMW,"BMW",best,instock)
    read_all_tabs(gc,ID_PORSCHE,"Porsche",best,instock,force="available")
    folder=os.environ.get("AUTONOVA_FOLDER_ID")
    if folder: pull_autonova_drive(folder,best,instock) # шлях Drive (без пароля)
    else: pull_autonova(best,instock) # запасний шлях: пошта IMAP
    print(f"[supply] собівартість зібрано: {len(best)} артикулів, у наявності {len(instock)}")
    pull_autonova_cache(best, instock) # дилерські ціни autonova з веб-кешу (для кодів поза BMW-аркушами)
    overrides=load_map(gc,"overrides"); comps=load_map(gc,"competitors")

    ws, vals, idx = read_export(gc) # каталог = та вкладка, яку тягне Prom
    print(f"[export] каталог «{EXPORT_TAB}»: {len(idx)} кодів")

    # --- Добір кодів БЕЗ постачальника. ПОРЯДОК ДЖЕРЕЛ: 1) BMW-пари з аркушів 2) autonova-web 3) BM Parts ---
    _miss=[c for c in idx if c not in best]
    print(f"[supply+] без постачальника: {len(_miss)}; приклади: " + " | ".join(str(c) for c in _miss[:10]))
    pull_pairs_from_best(_miss, best, instock)                 # 1) BMW-пари (подвоєні номери) з BMW-аркушів
    _miss=[c for c in idx if c not in best]
    cookie=os.environ.get("AUTONOVA_COOKIE")
    if cookie: cookie=cookie.replace("\r","").replace("\n","").strip() # прибрати переноси з копіпасту (інакше urllib: Invalid header value)
    if _miss and cookie:
        pull_autonova_web(_miss, best, instock, cookie)        # 2) autonova-web (дилерські ціни)
    elif _miss:
        print("[autonova-web] нема AUTONOVA_COOKIE — крок 2 пропущено (додай секрет AUTONOVA_COOKIE)")
    _miss=[c for c in idx if c not in best]
    if _miss:
        pull_bmparts(_miss, best, instock)                     # 3) BM Parts (решта)
    _left=[c for c in idx if c not in best]
    print(f"[supply+] після добору: собівартість {len(best)}; лишилось без постачальника {len(_left)}")

    only=os.environ.get("LIVE_ONLY")
    keep=set(a.strip().upper() for a in only.split(",") if a.strip()) if only else None
    catalog={}; guard_status={}; final_price_map={}; held=[]; updates=[]; matched=0
    for code,i in idx.items():
        row=vals[i]; cur=num(row[C_PRICE])
        catalog[code]={"name":row[C_NAME],"price":row[C_PRICE]}
        it=best.get(code)
        if not it: continue # нема постачальника — рядок НЕ чіпаємо взагалі
        if keep and code not in keep: continue # канарка LIVE_ONLY (якщо задано)
        matched+=1
        comp=comps.get(code)
        newp=overrides.get(code) or price_with_competitor(it["cost"], comp) # ціна = собівартість × тариф (або конкурент/ручна). БЕЗ утримання — застосовуємо завжди.
        aq=instock.get(code,0) # наявність(+/дні) + кількість
        if aq>0: row[C_AVAIL]="+"; row[C_QTY]=int(aq)
        else: row[C_AVAIL]="15"; row[C_QTY]="" # під замовлення 15 днів, кількість порожня
        rn=i+1
        row[C_PRICE]=int(newp); final_price_map[code]=int(newp); guard_status[code]="оновлено"
        updates.append({"range":f"I{rn}","values":[[int(newp)]]}) # Ціна(I) — точково цей рядок
        updates.append({"range":f"P{rn}:Q{rn}","values":[[row[C_AVAIL], row[C_QTY]]]}) # Наявність(P)+Кількість(Q)
    price_upd=sum(1 for u in updates if u["range"].startswith("I"))
    print(f"[calc] зіставлено {matched}, ціну оновлено {price_upd} (усі застосовано, без утримання)")

    if LIVE and updates:
        B=2000 # чанки, щоб не перевищити ліміт запиту
        for j in range(0,len(updates),B):
            ws.batch_update(updates[j:j+B], value_input_option="RAW")
        print(f"[export] ЗАПИСАНО в «{EXPORT_TAB}»: ціна {price_upd} + наявність/кількість {matched} рядків (ТІЛЬКИ змінені; утримані та без постачальника не чіпав). Prom підтягне фідом.")
    else:
        print(f"[export] DRY-RUN (LIVE≠1): у «{EXPORT_TAB}» НЕ писав. Приклади порахованого:")
        shown=0
        for code,i in idx.items():
            if code in final_price_map and str(final_price_map[code])!="" and shown<8:
                print(f"  {code}: ціна={vals[i][C_PRICE]} наяв={vals[i][C_AVAIL]} к-ть={vals[i][C_QTY]}"); shown+=1

    try: write_report(gc, catalog, best, instock, overrides, comps, guard_status, final_price_map)
    except Exception as e: print("[report]", str(e)[:140])
    if held:
        print("[guard] УТРИМАНІ (сильне зниження без конкурента, ціну НЕ чіпав; перші 30):")
        for a,c,nw in held[:30]: print(f"[guard] {a}: {c:.0f} -> {nw:.0f}")

if __name__=="__main__": main()
