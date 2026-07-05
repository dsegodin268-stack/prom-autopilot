import openpyxl
from engine.catalog import norm
def _avail(sheet, qty):
    s=str(sheet).lower()
    if any(k in s for k in ["чека","замов","2-3","2–3","заказ","підзам"]): return "під замовлення (2-3 дні)"
    if "наяв" in s: return "в наявності"
    try: return "в наявності" if float(qty)>0 else "під замовлення"
    except: return "під замовлення"
def parse_excel(path, columns, brand, availability=None, sheets=None, header_rows=0):
    ac,nc,qc,cc=columns.get("article",0),columns.get("name",1),columns.get("qty",2),columns.get("cost",3)
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True); out=[]
    for sh in (sheets or wb.sheetnames):
        for i,r in enumerate(wb[sh].iter_rows(values_only=True)):
            if i<header_rows or not r or len(r)<=max(ac,nc,qc,cc): continue
            art=r[ac]
            if art is None: continue
            try: cost=float(r[cc])
            except: continue
            if cost<=0: continue
            out.append({"article":norm(art),"name":str(r[nc] or ""),"qty":r[qc],
                        "cost":cost,"availability":availability or _avail(sh, r[qc]),"brand":brand})
    wb.close(); return out
